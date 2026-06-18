"""
correr_pruebas_excel.py - Automatizacion de pruebas K-QGMIP para DatosPruebas2026_1.xlsx

Lee las hojas de bitacora de la profesora, ejecuta KQNodes y KGeometricSIA
para k in {2, 3, 4, 5} en cada subsistema definido por Alcance/Mecanismo,
rellena las celdas de Particion/Perdida/Tiempo y genera las tres graficas
requeridas (A: tiempo vs. tamano, B: perdida comparativa, C: variacion EMD).

Uso:
    python3 correr_pruebas_excel.py
    python3 correr_pruebas_excel.py --excel path/al/archivo.xlsx
    python3 correr_pruebas_excel.py --solo-N 10        (solo hoja 10A)
    python3 correr_pruebas_excel.py --max-pruebas 5    (primeras 5 pruebas por hoja)
    python3 correr_pruebas_excel.py --sin-geo           (omite KGeomSIA)

Universidad de Caldas - Proyecto K-QGMIP - 2026-1
"""

import os
import sys
import json
import subprocess
import time
import argparse
import textwrap
from pathlib import Path
from copy import copy

# Forzar stdout en UTF-8 en Windows para evitar UnicodeEncodeError
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", line_buffering=True)

import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# ==============================================================================
# RUTAS Y CONFIGURACION
# ==============================================================================

SCRIPT_DIR  = Path(__file__).parent.resolve()
QNODES_ROOT = str(SCRIPT_DIR / "QNodes")
GEO_M2      = str(SCRIPT_DIR / "GeoMIP" / "src" /
                  "Method2_Dynamic_Programming_Reformulation")
SAMPLES_DIR = str(SCRIPT_DIR / "QNodes" / "src" / ".samples")
OUTPUT_DIR  = SCRIPT_DIR / "outputs" / "pruebas_excel"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Mapa: nombre de hoja -> CSV disponible en .samples/
HOJA_A_CSV = {
    "10A-Elementos":   os.path.join(SAMPLES_DIR, "N10A.csv"),
    "15B-Elementos":   os.path.join(SAMPLES_DIR, "N15B.csv"),
    "20A-Elementos":   None,   # TPM sintetica generada en subprocess
    "22A-Elementos":   None,
    "25A-Elementos ":  None,   # ojo: tiene espacio final en el nombre del xlsx
}

# Mapa inverso: nombre_hoja -> N (numero de nodos)
HOJA_A_N = {
    "10A-Elementos":  10,
    "15B-Elementos":  15,
    "20A-Elementos":  20,
    "22A-Elementos":  22,
    "25A-Elementos ": 25,
}

# Umbral a partir del cual KGeomSIA se omite automaticamente (2^N muy grande)
N_MAX_GEOMIP    = 10    # N=10 -> 1024 estados, manejable
TIMEOUT_KQNODES = 300   # segundos por llamada
TIMEOUT_KGEOMIP = 90    # segundos por llamada
# Para N>N_MAX_GEOMIP y N<=20: generamos TPM sintetica y usamos solo KQNodes
N_MAX_SINTETICA = 20    # N=22 requiere ~740 MB de RAM; N=25 ~6.7 GB

# Colores para Excel
FILL_HEADER    = PatternFill("solid", fgColor="2E75B6")   # azul encabezado
FILL_QNODES    = PatternFill("solid", fgColor="D9E8F5")   # azul claro
FILL_GEO       = PatternFill("solid", fgColor="D9F0E0")   # verde claro
FILL_SINTETICA = PatternFill("solid", fgColor="FFF2CC")   # amarillo suave
FILL_NA        = PatternFill("solid", fgColor="F2F2F2")   # gris claro
FONT_BOLD      = Font(bold=True)
ALIGN_CENTER   = Alignment(horizontal="center", vertical="center")

# ==============================================================================
# CONSTANTES K-QGMIP
# ==============================================================================

K_VALORES = [2, 3, 4, 5]

# Columna de inicio para cada k en la hoja (1-indexado, col 1 = #Prueba)
# Estructura: [QNodes_Particion | QNodes_Perdida | QNodes_Tiempo |
#              Geo_Particion    | Geo_Perdida    | Geo_Tiempo   ] x 4 bloques
K_BASE_COL = {2: 4, 3: 10, 4: 16, 5: 22}   # columna del primer campo de k

OFFSET_QNODES = 0   # QNodes: K_BASE_COL[k] + 0,1,2
OFFSET_GEO    = 3   # Geometric: K_BASE_COL[k] + 3,4,5

# ==============================================================================
# STUBS PARA IMPORTACIONES OPCIONALES
# ==============================================================================

_STUBS = """
import sys, types, numpy as np
def _stub(name, **attrs):
    m = types.ModuleType(name); m.__path__ = []
    for k, v in attrs.items(): setattr(m, k, v)
    sys.modules[name] = m; return m
_pyemd = _stub("pyemd"); _pyemd.emd = lambda u,v,c: float(np.sum(np.abs(u-v)))
_pi = _stub("pyinstrument")
class _FP:
    def __init__(self,**k): pass
    def start(self): pass
    def stop(self): return self
    def output(self, renderer=None): return ""
_pi.Profiler = _FP
_stub("pyinstrument.renderers").HTMLRenderer = type("H",(),{"__init__":lambda s,**k:None})
_stub("pyttsx3"); _stub("pyttsx3.engine",Engine=type("E",(),{})); _stub("pyttsx3.voice",Voice=type("V",(),{}))
_col = _stub("colorama"); _col.init=lambda**k:None
_col.Fore=type("F",(),{"__getattr__":lambda s,k:""})(); _col.Style=type("S",(),{"__getattr__":lambda s,k:""})()
_stub("colorama.ansi")
"""

# ==============================================================================
# UTILIDADES
# ==============================================================================

def letras_a_mascara(sistema: str, letras: str) -> str:
    """
    Convierte una cadena de letras a mascara binaria segun el sistema.

    Ejemplo:
        sistema = "ABCDEFGHIJ"
        letras  = "ACEGI"
        retorna = "1010101010"   (A=pos0, C=pos2, E=pos4, G=pos6, I=pos8)
    """
    return "".join("1" if c in letras else "0" for c in sistema)


def estado_a_str(raw) -> str:
    """
    Asegura que el estado_inicial leido del Excel sea una cadena de '0'/'1'.
    Excel puede almacenarlo como entero (e.g., 1000000000) o string.
    """
    s = str(raw).strip()
    if all(c in "01" for c in s):
        return s
    # Fallback: padding de ceros a la izquierda si necesario
    return s


def log(msg: str, nivel: str = "INFO"):
    """Impresion de estado con prefijos ASCII seguros para Windows."""
    prefijos = {
        "INFO": "[*]",
        "OK":   "[OK]",
        "WARN": "[!]",
        "ERR":  "[ERR]",
        "SKIP": "[->]",
    }
    prefijo = prefijos.get(nivel, "[*]")
    print(f"  {prefijo} {msg}", flush=True)


# ==============================================================================
# GENERACION DE CODIGO SUBPROCESS
# ==============================================================================

def _codigo_kqnodes(csv_path, estado, condicion, alcance, mecanismo, k,
                    N_sintetica=None):
    """
    Genera codigo Python para ejecutar KQNodes en un subprocess aislado.

    Si csv_path es None y N_sintetica esta definido, genera una TPM sintetica
    aleatoria de forma (2^N, N) directamente en el subprocess (en memoria).
    """
    carga_tpm = ""
    if csv_path and os.path.exists(csv_path):
        carga_tpm = f"tpm = np.loadtxt({csv_path!r}, delimiter=',')"
    elif N_sintetica is not None:
        carga_tpm = textwrap.dedent(f"""
        # TPM sintetica: N={N_sintetica}, seed=42 (reproducible)
        rng = np.random.default_rng(42)
        tpm = rng.random((2**{N_sintetica}, {N_sintetica}))
        """).strip()
    else:
        raise ValueError("Se requiere csv_path o N_sintetica.")

    return _STUBS + f"""
import os, importlib.machinery, time, json
QNODES_ROOT = {QNODES_ROOT!r}
sys.path.insert(0, QNODES_ROOT)

from src.middlewares.slogger import SafeLogger
SafeLogger.critic = lambda self, msg: None   # silenciar logs

import importlib.machinery
loader = importlib.machinery.SourceFileLoader(
    "k_qnodes_mod",
    os.path.join(QNODES_ROOT, "src", "strategies", "k_qnodes.py")
)
mod     = loader.load_module()
KQNodes = mod.KQNodes

try:
    {carga_tpm}
    kqn = KQNodes(tpm)
    t0  = time.perf_counter()
    res = kqn.aplicar_estrategia({estado!r}, {condicion!r}, {alcance!r}, {mecanismo!r}, k={k})
    t1  = time.perf_counter()
    print("RESULTADO|" + json.dumps({{
        "perdida":   float(res.perdida),
        "particion": str(res.particion),
        "tiempo":    float(t1 - t0),
        "algoritmo": "KQNodes"
    }}), flush=True)
except Exception as exc:
    import traceback
    print("ERROR|" + str(exc), flush=True)
    traceback.print_exc(file=sys.stderr)
    sys.exit(1)
"""


def _codigo_kgeomip(csv_path, estado, condicion, alcance, mecanismo, k):
    """
    Genera codigo Python para ejecutar KGeometricSIA en un subprocess aislado.
    Solo se llama cuando csv_path existe (sistemas reales <= N=10 en la practica).
    """
    if not (csv_path and os.path.exists(csv_path)):
        raise ValueError("KGeomSIA requiere CSV real.")

    return _STUBS + f"""
import os, tempfile, time, json
from pathlib import Path
QNODES_ROOT = {QNODES_ROOT!r}
GEO_M2      = {GEO_M2!r}
sys.path.insert(0, QNODES_ROOT)
sys.path.insert(0, GEO_M2)

from src.controllers.strategies.k_geometric import KGeometricSIA
from src.middlewares.slogger import SafeLogger

class _MM:
    def __init__(self, ei):
        self.estado_inicial     = ei
        self.pagina_red_muestra = "A"
    @property
    def pagina(self): return "A"
    @property
    def output_dir(self):
        p = Path(tempfile.gettempdir()) / "kgeomip_pruebas"; p.mkdir(exist_ok=True); return p
    @property
    def tpm_filename(self): return None

try:
    tpm = np.loadtxt({csv_path!r}, delimiter=',')
    obj = KGeometricSIA.__new__(KGeometricSIA)
    obj.sia_gestor          = _MM({estado!r})
    obj.sia_logger          = type("L",(),{{"critic":lambda s,m:None,"warn":lambda s,m:None,"debug":lambda s,m:None}})()
    obj.logger              = SafeLogger("kgeomip_pruebas")
    obj.k                   = {k}
    obj.etiquetas           = [[], []]
    obj.tabla_transiciones  = {{}}
    obj.vertices            = set()
    obj.tabla               = {{}}
    obj.memoria_particiones = {{}}
    obj.sia_tiempo_inicio   = 0.0

    t0  = time.perf_counter()
    res = obj.aplicar_estrategia({condicion!r}, {alcance!r}, {mecanismo!r}, tpm, k={k})
    t1  = time.perf_counter()
    print("RESULTADO|" + json.dumps({{
        "perdida":   float(res.perdida),
        "particion": str(res.particion),
        "tiempo":    float(t1 - t0),
        "algoritmo": "KGeometricSIA"
    }}), flush=True)
except Exception as exc:
    import traceback
    print("ERROR|" + str(exc), flush=True)
    traceback.print_exc(file=sys.stderr)
    sys.exit(1)
"""


# ==============================================================================
# EJECUTORES SUBPROCESS
# ==============================================================================

def _ejecutar_subprocess(codigo: str, timeout: int) -> dict | None:
    """
    Ejecuta codigo Python en un subprocess aislado y retorna el dict
    del resultado (de la linea RESULTADO|{json}), o None si falla/timeout.
    """
    try:
        proc = subprocess.run(
            [sys.executable, "-c", codigo],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=timeout,
        )
        for linea in proc.stdout.splitlines():
            if linea.startswith("RESULTADO|"):
                return json.loads(linea[len("RESULTADO|"):])
        return None
    except subprocess.TimeoutExpired:
        return None
    except Exception:
        return None


def ejecutar_kqnodes(csv_path, estado, condicion, alcance, mecanismo, k,
                     N_sintetica=None) -> dict:
    """
    Ejecuta KQNodes y retorna {'particion', 'perdida', 'tiempo'} o marcadores N/A.
    """
    codigo = _codigo_kqnodes(csv_path, estado, condicion, alcance, mecanismo, k,
                              N_sintetica)
    res = _ejecutar_subprocess(codigo, TIMEOUT_KQNODES)
    if res is None:
        return {"particion": "TIMEOUT", "perdida": None, "tiempo": None}
    return res


def ejecutar_kgeomip(csv_path, estado, condicion, alcance, mecanismo, k) -> dict:
    """
    Ejecuta KGeomSIA y retorna {'particion', 'perdida', 'tiempo'} o marcadores N/A.
    """
    codigo = _codigo_kgeomip(csv_path, estado, condicion, alcance, mecanismo, k)
    res = _ejecutar_subprocess(codigo, TIMEOUT_KGEOMIP)
    if res is None:
        return {"particion": "TIMEOUT", "perdida": None, "tiempo": None}
    return res


# ==============================================================================
# LECTURA DEL EXCEL
# ==============================================================================

def leer_hoja(ws, nombre: str) -> dict:
    """
    Extrae metadata y pruebas de una hoja de bitacora.

    Retorna:
        {
          'nombre': str,
          'N': int,
          'sistema': str,        # e.g., "ABCDEFGHIJ"
          'estado': str,         # e.g., "1000000000"
          'csv': str|None,       # path al CSV real o None
          'pruebas': [{'num', 'alcance_letras', 'mecanismo_letras',
                       'alcance_bin', 'mecanismo_bin'}, ...]
        }
    """
    estado_raw = ws.cell(1, 2).value
    sistema    = str(ws.cell(2, 2).value or "").strip()
    N          = len(sistema)

    estado = estado_a_str(estado_raw)
    if len(estado) != N:
        # Podria ser un int que perdio ceros iniciales, reconstruir
        estado = estado.zfill(N)

    csv_path = HOJA_A_CSV.get(nombre)

    pruebas = []
    for row_idx in range(6, ws.max_row + 1):
        num_cell = ws.cell(row_idx, 1).value
        if num_cell is None:
            break
        alc_letras = str(ws.cell(row_idx, 2).value or "").strip()
        mec_letras = str(ws.cell(row_idx, 3).value or "").strip()
        if not alc_letras or not mec_letras:
            continue
        pruebas.append({
            "num":              int(num_cell),
            "fila":             row_idx,
            "alcance_letras":   alc_letras,
            "mecanismo_letras": mec_letras,
            "alcance_bin":      letras_a_mascara(sistema, alc_letras),
            "mecanismo_bin":    letras_a_mascara(sistema, mec_letras),
        })

    return {
        "nombre":  nombre,
        "N":       N,
        "sistema": sistema,
        "estado":  estado,
        "csv":     csv_path,
        "pruebas": pruebas,
    }


# ==============================================================================
# ESCRITURA DE RESULTADOS EN EXCEL
# ==============================================================================

def _escribir_celda(ws, fila, col, valor, fill=None, bold=False, fmt=None):
    """Escribe un valor en una celda con formato opcional."""
    c = ws.cell(fila, col, valor)
    c.alignment = ALIGN_CENTER
    if fill:
        c.fill = fill
    if bold:
        c.font = Font(bold=True)
    if fmt:
        c.number_format = fmt


def escribir_resultados_en_hoja(ws, prueba_fila: int, k: int,
                                 res_qn: dict, res_geo: dict,
                                 es_sintetica: bool = False):
    """
    Escribe los resultados de una prueba para un k dado en la hoja.

    Columnas de destino (segun estructura del Excel):
      k=2: C4-C9   k=3: C10-C15   k=4: C16-C21   k=5: C22-C27
    """
    base   = K_BASE_COL[k]
    fill_q = FILL_SINTETICA if es_sintetica else FILL_QNODES
    fill_g = FILL_NA        if es_sintetica else FILL_GEO

    # -- KQNodes (cols base, base+1, base+2) --
    _escribir_celda(ws, prueba_fila, base + 0,
                    str(res_qn.get("particion", "ERR"))[:80], fill=fill_q)
    perdida_qn = res_qn.get("perdida")
    _escribir_celda(ws, prueba_fila, base + 1,
                    round(perdida_qn, 6) if perdida_qn is not None else "N/A",
                    fill=fill_q, fmt="0.000000")
    tiempo_qn = res_qn.get("tiempo")
    _escribir_celda(ws, prueba_fila, base + 2,
                    round(tiempo_qn, 6) if tiempo_qn is not None else "N/A",
                    fill=fill_q, fmt="0.000000")

    # -- KGeomSIA (cols base+3, base+4, base+5) --
    _escribir_celda(ws, prueba_fila, base + 3,
                    str(res_geo.get("particion", "N/A"))[:80], fill=fill_g)
    perdida_g = res_geo.get("perdida")
    _escribir_celda(ws, prueba_fila, base + 4,
                    round(perdida_g, 6) if perdida_g is not None else "N/A",
                    fill=fill_g, fmt="0.000000")
    tiempo_g = res_geo.get("tiempo")
    _escribir_celda(ws, prueba_fila, base + 5,
                    round(tiempo_g, 6) if tiempo_g is not None else "N/A",
                    fill=fill_g, fmt="0.000000")


# ==============================================================================
# NUCLEO: PROCESAR TODAS LAS HOJAS
# ==============================================================================

def procesar_excel(excel_path: str, max_pruebas: int = None,
                   solo_N: int = None, sin_geo: bool = False) -> tuple[list, Path]:
    """
    Lee el Excel, ejecuta los algoritmos y escribe resultados.

    Retorna:
        (resultados, ruta_excel_salida)

    resultados: lista de dicts con todos los datos numericos recolectados,
                usados para generar las graficas.
    """
    SEP_DOBLE  = "=" * 64
    SEP_SIMPLE = "-" * 64

    print(f"\n{SEP_DOBLE}", flush=True)
    print("  K-QGMIP - Automatizacion de pruebas", flush=True)
    print(f"  Excel: {excel_path}", flush=True)
    print(f"{SEP_DOBLE}\n", flush=True)

    wb_src  = openpyxl.load_workbook(excel_path, data_only=True)
    wb_dest = openpyxl.load_workbook(excel_path)   # copia para escribir

    nombre_salida = Path(excel_path).stem + "_RESULTADOS.xlsx"
    ruta_salida   = OUTPUT_DIR / nombre_salida

    resultados = []   # lista de filas para las graficas
    hojas_a_procesar = [n for n in HOJA_A_CSV if n in wb_src.sheetnames]

    for nombre_hoja in hojas_a_procesar:
        info = leer_hoja(wb_src[nombre_hoja], nombre_hoja)
        N    = info["N"]

        if solo_N is not None and N != solo_N:
            log(f"Omitiendo {nombre_hoja} (N={N}, --solo-N={solo_N})", "SKIP")
            continue

        print(f"\n{SEP_SIMPLE}", flush=True)
        print(f"  Hoja: {nombre_hoja}  |  N={N}  |  Sistema: {info['sistema']}",
              flush=True)
        print(f"  Estado inicial: {info['estado']}", flush=True)
        csv_disp = info['csv'] if info['csv'] else "(se generara TPM sintetica)"
        print(f"  CSV real: {csv_disp}", flush=True)
        print(f"  Pruebas: {len(info['pruebas'])}", flush=True)
        print(SEP_SIMPLE, flush=True)

        # Determinar si es sintetica o real
        es_sintetica = (info["csv"] is None or not os.path.exists(info["csv"] or ""))
        usar_geo     = (not sin_geo) and (N <= N_MAX_GEOMIP) and not es_sintetica

        if N > N_MAX_SINTETICA and es_sintetica:
            gb_est = 2 ** N * N * 8 / 1e9
            log(f"N={N} > {N_MAX_SINTETICA}: TPM sintetica omitida "
                f"(requeriria ~{gb_est:.1f} GB RAM). Hoja marcada N/A.", "WARN")
            ws_dest = wb_dest[nombre_hoja]
            for p in info["pruebas"][:max_pruebas or len(info["pruebas"])]:
                for k in K_VALORES:
                    _na = {"particion": "N/A (N muy grande)",
                           "perdida": None, "tiempo": None}
                    escribir_resultados_en_hoja(ws_dest, p["fila"], k, _na, _na, True)
            continue

        pruebas_sel = info["pruebas"][:max_pruebas] if max_pruebas else info["pruebas"]
        ws_dest     = wb_dest[nombre_hoja]
        condicion   = "1" * N   # todas las variables activas (sin background)

        for idx, prueba in enumerate(pruebas_sel):
            num   = prueba["num"]
            alc   = prueba["alcance_bin"]
            mec   = prueba["mecanismo_bin"]
            fila  = prueba["fila"]
            n_alc = alc.count("1")
            n_mec = mec.count("1")

            alc_short = prueba['alcance_letras'][:20]
            mec_short = prueba['mecanismo_letras'][:20]
            print(f"\n  P{num:>3} | Alc={alc_short} ({n_alc}n)"
                  f" | Mec={mec_short} ({n_mec}n)", flush=True)

            for k in K_VALORES:
                # -- KQNodes --
                res_qn = ejecutar_kqnodes(
                    csv_path=info["csv"],
                    estado=info["estado"],
                    condicion=condicion,
                    alcance=alc,
                    mecanismo=mec,
                    k=k,
                    N_sintetica=N if es_sintetica else None,
                )
                estado_qn = "OK" if res_qn.get("perdida") is not None else "FAIL"
                phi_qn = res_qn.get('perdida', 'N/A')
                t_qn   = res_qn.get('tiempo', 'N/A')
                log(f"k={k} KQNodes   : phi={phi_qn}  t={t_qn}  [{estado_qn}]",
                    "OK" if estado_qn == "OK" else "WARN")

                # -- KGeomSIA --
                if usar_geo:
                    res_geo = ejecutar_kgeomip(
                        csv_path=info["csv"],
                        estado=info["estado"],
                        condicion=condicion,
                        alcance=alc,
                        mecanismo=mec,
                        k=k,
                    )
                    estado_geo = "OK" if res_geo.get("perdida") is not None else "FAIL"
                    phi_geo = res_geo.get('perdida', 'N/A')
                    t_geo   = res_geo.get('tiempo', 'N/A')
                    log(f"k={k} KGeomSIA  : phi={phi_geo}  t={t_geo}  [{estado_geo}]",
                        "OK" if estado_geo == "OK" else "WARN")
                else:
                    if sin_geo:
                        motivo = "--sin-geo activo"
                    elif es_sintetica:
                        motivo = "TPM sintetica"
                    else:
                        motivo = f"N={N} > limite ({N_MAX_GEOMIP})"
                    res_geo = {"particion": f"N/A ({motivo})",
                               "perdida": None, "tiempo": None}
                    log(f"k={k} KGeomSIA  : omitido - {motivo}", "SKIP")

                # -- Escribir en hoja destino --
                escribir_resultados_en_hoja(ws_dest, fila, k, res_qn, res_geo,
                                            es_sintetica)

                # -- Acumular para graficas --
                resultados.append({
                    "hoja":          nombre_hoja,
                    "N":             N,
                    "prueba":        num,
                    "k":             k,
                    "sintetica":     es_sintetica,
                    "n_alcance":     n_alc,
                    "n_mecanismo":   n_mec,
                    "qn_perdida":    res_qn.get("perdida"),
                    "qn_tiempo":     res_qn.get("tiempo"),
                    "qn_particion":  res_qn.get("particion", ""),
                    "geo_perdida":   res_geo.get("perdida"),
                    "geo_tiempo":    res_geo.get("tiempo"),
                    "geo_particion": res_geo.get("particion", ""),
                })

        # Guardar progreso despues de cada hoja
        wb_dest.save(ruta_salida)
        log(f"Guardado parcial -> {ruta_salida.name}", "OK")

    wb_dest.save(ruta_salida)
    print(f"\n{SEP_DOBLE}", flush=True)
    print(f"  [OK] Excel de resultados: {ruta_salida}", flush=True)
    print(f"{SEP_DOBLE}\n", flush=True)
    return resultados, ruta_salida


# ==============================================================================
# GENERACION DE GRAFICAS (A, B, C)
# ==============================================================================

COLOR_QN  = "#2E75B6"   # azul KQNodes
COLOR_GEO = "#37A84E"   # verde KGeomSIA
COLOR_K   = {2: "#E74C3C", 3: "#F39C12", 4: "#8E44AD", 5: "#1ABC9C"}


def _validar_datos_num(resultados: list) -> list:
    """Filtra resultados con al menos qn_tiempo valido."""
    return [r for r in resultados if r["qn_tiempo"] is not None]


def grafica_A_tiempo_por_sistema(resultados: list) -> Path:
    """
    Grafica A: Tiempo de ejecucion medio por tamano de sistema.

    X-axis : tamanos de sistema ordenados [10, 15, 20, 22, 25]
    Y-axis : tiempo promedio (segundos, escala log)
    Lineas : una por algoritmo x k
    """
    datos = _validar_datos_num(resultados)
    if not datos:
        log("Sin datos para Grafica A.", "WARN")
        return None

    N_orden = [10, 15, 20, 22, 25]
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    fig.suptitle("Grafica A - Tiempo de Ejecucion por Tamano de Sistema",
                 fontsize=14, fontweight="bold", y=1.02)

    # Panel izquierdo: tiempo por N promediado sobre k y pruebas
    ax = axes[0]
    for algo, col, marker in [("KQNodes", COLOR_QN, "o"), ("KGeomSIA", COLOR_GEO, "s")]:
        tiempos_N = {}
        for r in datos:
            t = r["qn_tiempo"] if algo == "KQNodes" else r["geo_tiempo"]
            if t is not None:
                tiempos_N.setdefault(r["N"], []).append(t)
        Ns     = [n for n in N_orden if n in tiempos_N]
        medias = [np.mean(tiempos_N[n]) for n in Ns]
        if Ns:
            ax.plot(Ns, medias, marker=marker, color=col, linewidth=2,
                    markersize=8, label=algo)
    ax.set_xlabel("Tamano del sistema (N nodos)", fontsize=11)
    ax.set_ylabel("Tiempo promedio (s)", fontsize=11)
    ax.set_yscale("log")
    ax.set_xticks(N_orden)
    ax.set_title("Promedio por sistema (escala log)")
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)

    # Panel derecho: desglose por k para KQNodes
    ax2 = axes[1]
    for k in K_VALORES:
        tiempos_N = {}
        for r in datos:
            if r["k"] == k and r["qn_tiempo"] is not None:
                tiempos_N.setdefault(r["N"], []).append(r["qn_tiempo"])
        Ns     = [n for n in N_orden if n in tiempos_N]
        medias = [np.mean(tiempos_N[n]) for n in Ns]
        if Ns:
            ax2.plot(Ns, medias, marker="o", color=COLOR_K[k], linewidth=2,
                     markersize=7, label=f"k={k}")
    ax2.set_xlabel("Tamano del sistema (N nodos)", fontsize=11)
    ax2.set_ylabel("Tiempo KQNodes (s)", fontsize=11)
    ax2.set_yscale("log")
    ax2.set_xticks(N_orden)
    ax2.set_title("KQNodes - desglose por k (escala log)")
    ax2.legend(fontsize=10)
    ax2.grid(True, alpha=0.3)

    plt.tight_layout()
    ruta = OUTPUT_DIR / "graficaA_tiempo_por_sistema.png"
    plt.savefig(ruta, dpi=150, bbox_inches="tight")
    plt.close()
    log(f"Grafica A guardada -> {ruta.name}", "OK")
    return ruta


def grafica_B_perdida_comparativa(resultados: list) -> Path:
    """
    Grafica B: Comparacion de perdida (EMD) entre QNodes y KGeomSIA.

    Un panel por k (k=2,3,4,5). Barras agrupadas por prueba.
    Muestra si ambos algoritmos encuentran la misma k-particion.
    """
    datos_ambos = [r for r in resultados
                   if r["qn_perdida"] is not None and r["geo_perdida"] is not None]

    if not datos_ambos:
        log("Sin datos comparativos (ambos algoritmos) para Grafica B.", "WARN")
        # Grafica alternativa: solo KQNodes, perdida por N y k
        datos = _validar_datos_num(resultados)
        if not datos:
            return None
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        fig.suptitle("Grafica B - Perdida (EMD) por Subsistema - KQNodes",
                     fontsize=14, fontweight="bold")
        for idx, k in enumerate(K_VALORES):
            ax = axes[idx // 2][idx % 2]
            dk = [r for r in datos if r["k"] == k]
            hojas_unicas = sorted(set(r["hoja"] for r in dk))
            x_labels, y_vals = [], []
            for h in hojas_unicas:
                vals = [r["qn_perdida"] for r in dk
                        if r["hoja"] == h and r["qn_perdida"] is not None]
                x_labels.append(h.replace("-Elementos", ""))
                y_vals.append(np.mean(vals) if vals else 0)
            ax.bar(x_labels, y_vals, color=COLOR_QN, alpha=0.8, edgecolor="white")
            ax.set_title(f"k = {k}", fontsize=12, fontweight="bold")
            ax.set_xlabel("Subsistema")
            ax.set_ylabel("Perdida EMD promedio")
            ax.grid(axis="y", alpha=0.3)
            ax.tick_params(axis="x", rotation=30)
        plt.tight_layout()
        ruta = OUTPUT_DIR / "graficaB_perdida_comparativa.png"
        plt.savefig(ruta, dpi=150, bbox_inches="tight")
        plt.close()
        log(f"Grafica B guardada -> {ruta.name}", "OK")
        return ruta

    # Si tenemos datos de ambos algoritmos
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle("Grafica B - Perdida EMD: QNodes vs. KGeomSIA por k-particion",
                 fontsize=14, fontweight="bold")

    for idx, k in enumerate(K_VALORES):
        ax = axes[idx // 2][idx % 2]
        dk = [r for r in datos_ambos if r["k"] == k]
        if not dk:
            ax.set_visible(False)
            continue
        nums   = [r["prueba"] for r in dk]
        q_vals = [r["qn_perdida"]  for r in dk]
        g_vals = [r["geo_perdida"] for r in dk]
        x = np.arange(len(nums))
        w = 0.38
        ax.bar(x - w/2, q_vals, w, label="KQNodes",   color=COLOR_QN,  alpha=0.85)
        ax.bar(x + w/2, g_vals, w, label="KGeomSIA",  color=COLOR_GEO, alpha=0.85)
        ax.set_xticks(x)
        ax.set_xticklabels([f"P{n}" for n in nums], fontsize=7, rotation=45)
        ax.set_title(f"k = {k}", fontsize=12, fontweight="bold")
        ax.set_xlabel("Prueba")
        ax.set_ylabel("Perdida EMD (phi)")
        ax.legend(fontsize=9)
        ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    ruta = OUTPUT_DIR / "graficaB_perdida_comparativa.png"
    plt.savefig(ruta, dpi=150, bbox_inches="tight")
    plt.close()
    log(f"Grafica B guardada -> {ruta.name}", "OK")
    return ruta


def grafica_C_variacion_emd(resultados: list) -> Path:
    """
    Grafica C: Variacion de perdida usando QNodes como referencia (= 0).

    Muestra delta = geo_perdida - qn_perdida para cada prueba.
    delta > 0 -> KGeomSIA encontro peor particion (mas perdida).
    delta < 0 -> KGeomSIA encontro mejor particion.
    delta = 0 -> coinciden.

    Un panel por k. Solo pruebas donde ambos algoritmos retornaron resultado.
    """
    datos_ambos = [r for r in resultados
                   if r["qn_perdida"] is not None and r["geo_perdida"] is not None]

    if not datos_ambos:
        log("Sin datos comparativos para Grafica C - dibujando distribucion KQNodes.",
            "WARN")
        datos = _validar_datos_num(resultados)
        if not datos:
            return None
        fig, axes = plt.subplots(2, 2, figsize=(14, 9))
        fig.suptitle("Grafica C - Distribucion de Perdida KQNodes\n"
                     "(KGeomSIA no disponible para estos sistemas)",
                     fontsize=13, fontweight="bold")
        for idx, k in enumerate(K_VALORES):
            ax = axes[idx // 2][idx % 2]
            dk = [r for r in datos
                  if r["k"] == k and r["qn_perdida"] is not None]
            if not dk:
                ax.set_visible(False)
                continue
            vals = [r["qn_perdida"] for r in dk]
            ax.hist(vals, bins=20, color=COLOR_QN, alpha=0.8, edgecolor="white")
            media = np.mean(vals)
            ax.axvline(media, color="#E74C3C", linewidth=2,
                       linestyle="--", label=f"Media = {media:.3f}")
            ax.set_title(f"k = {k} - Distribucion phi (KQNodes)",
                         fontsize=11, fontweight="bold")
            ax.set_xlabel("Perdida EMD (phi)")
            ax.set_ylabel("Frecuencia")
            ax.legend(fontsize=9)
            ax.grid(alpha=0.3)
        plt.tight_layout()
        ruta = OUTPUT_DIR / "graficaC_variacion_emd.png"
        plt.savefig(ruta, dpi=150, bbox_inches="tight")
        plt.close()
        log(f"Grafica C guardada -> {ruta.name}", "OK")
        return ruta

    fig, axes = plt.subplots(2, 2, figsize=(14, 9))
    fig.suptitle("Grafica C - Variacion de Perdida EMD\n"
                 "(KQNodes como referencia = 0, delta = KGeomSIA - KQNodes)",
                 fontsize=13, fontweight="bold")

    for idx, k in enumerate(K_VALORES):
        ax = axes[idx // 2][idx % 2]
        dk = [r for r in datos_ambos if r["k"] == k]
        if not dk:
            ax.set_visible(False)
            continue
        nums   = [r["prueba"] for r in dk]
        deltas = [r["geo_perdida"] - r["qn_perdida"] for r in dk]
        colors = [COLOR_GEO if d >= 0 else COLOR_QN for d in deltas]
        x = np.arange(len(nums))
        ax.bar(x, deltas, color=colors, alpha=0.85, edgecolor="white")
        ax.axhline(0, color="black", linewidth=1.2, linestyle="-")
        ax.set_xticks(x)
        ax.set_xticklabels([f"P{n}" for n in nums], fontsize=7, rotation=45)
        ax.set_xlabel("Prueba")
        ax.set_ylabel("delta = phi(Geo) - phi(QNodes)")
        n_pos = sum(1 for d in deltas if d > 0)
        n_neg = sum(1 for d in deltas if d < 0)
        n_zer = sum(1 for d in deltas if d == 0)
        ax.set_title(
            f"k={k}  |  Geo>QN: {n_pos}  Geo<QN: {n_neg}  Iguales: {n_zer}",
            fontsize=10, fontweight="bold"
        )
        ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    ruta = OUTPUT_DIR / "graficaC_variacion_emd.png"
    plt.savefig(ruta, dpi=150, bbox_inches="tight")
    plt.close()
    log(f"Grafica C guardada -> {ruta.name}", "OK")
    return ruta


def generar_graficas(resultados: list):
    """Genera las tres graficas requeridas y retorna sus rutas."""
    SEP = "-" * 64
    print(f"\n{SEP}", flush=True)
    print("  Generando graficas A, B, C...", flush=True)
    print(SEP, flush=True)
    rA = grafica_A_tiempo_por_sistema(resultados)
    rB = grafica_B_perdida_comparativa(resultados)
    rC = grafica_C_variacion_emd(resultados)
    return rA, rB, rC


# ==============================================================================
# RESUMEN FINAL
# ==============================================================================

def imprimir_resumen(resultados: list, ruta_excel: Path, rA, rB, rC):
    """Muestra estadisticas globales al finalizar."""
    SEP = "=" * 64
    print(f"\n{SEP}", flush=True)
    print("  RESUMEN DE EJECUCION", flush=True)
    print(SEP, flush=True)

    total  = len(resultados)
    ok_qn  = sum(1 for r in resultados if r["qn_perdida"]  is not None)
    ok_geo = sum(1 for r in resultados if r["geo_perdida"] is not None)

    print(f"  Total ejecuciones registradas : {total}", flush=True)
    print(f"  KQNodes con resultado valido  : {ok_qn}/{total}", flush=True)
    print(f"  KGeomSIA con resultado valido : {ok_geo}/{total}", flush=True)

    print(f"\n  Por tamano de sistema:", flush=True)
    for N in [10, 15, 20, 22, 25]:
        subset = [r for r in resultados if r["N"] == N]
        if not subset:
            continue
        qn_t  = [r["qn_tiempo"]  for r in subset if r["qn_tiempo"]  is not None]
        qn_p  = [r["qn_perdida"] for r in subset if r["qn_perdida"] is not None]
        tipo  = "[!] TPM sintetica" if subset[0]["sintetica"] else "[OK] CSV real"
        if qn_t and qn_p:
            print(f"    N={N:>2} ({tipo}): {len(subset)} exec | "
                  f"KQN t_avg={np.mean(qn_t):.3f}s  phi_avg={np.mean(qn_p):.3f}",
                  flush=True)
        else:
            print(f"    N={N:>2} ({tipo}): {len(subset)} exec | sin datos",
                  flush=True)

    print(f"\n  Archivos generados:", flush=True)
    print(f"    - {ruta_excel}", flush=True)
    if rA:
        print(f"    - {rA}", flush=True)
    if rB:
        print(f"    - {rB}", flush=True)
    if rC:
        print(f"    - {rC}", flush=True)
    print(f"\n{SEP}\n", flush=True)


# ==============================================================================
# PUNTO DE ENTRADA
# ==============================================================================

def _arg_parser():
    parser = argparse.ArgumentParser(
        description="Ejecuta K-QGMIP sobre las pruebas del Excel de la profesora.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""
        Ejemplos:
          python3 correr_pruebas_excel.py
          python3 correr_pruebas_excel.py --excel ../DatosPruebas2026_1.xlsx
          python3 correr_pruebas_excel.py --solo-N 10 --max-pruebas 5
          python3 correr_pruebas_excel.py --sin-geo
        """)
    )
    parser.add_argument(
        "--excel",
        default=None,
        help="Ruta al archivo DatosPruebas2026_1.xlsx "
             "(por defecto busca en la carpeta del script y doc_biblioteca/docProfesora/)."
    )
    parser.add_argument(
        "--max-pruebas",
        type=int,
        default=None,
        metavar="N",
        help="Limita el numero de pruebas por hoja (util para verificacion rapida)."
    )
    parser.add_argument(
        "--solo-N",
        type=int,
        default=None,
        metavar="N",
        help="Procesa solo la hoja de N nodos (10, 15, 20, 22 o 25)."
    )
    parser.add_argument(
        "--sin-geo",
        action="store_true",
        help="Omite KGeometricSIA completamente (modo rapido)."
    )
    return parser


def _encontrar_excel_automatico() -> str:
    """Busca el Excel en ubicaciones tipicas del proyecto."""
    candidatos = [
        SCRIPT_DIR / "DatosPruebas2026_1.xlsx",
        SCRIPT_DIR.parent / "DatosPruebas2026_1.xlsx",
        SCRIPT_DIR / "docs" / "DatosPruebas2026_1.xlsx",
        SCRIPT_DIR / "docs" / "doc_biblioteca" / "docProfesora" / "DatosPruebas2026_1.xlsx",
    ]
    for ruta in candidatos:
        if ruta.exists():
            return str(ruta)
    return None


def main():
    parser = _arg_parser()
    args   = parser.parse_args()

    # Resolver ruta del Excel
    excel_path = args.excel
    if excel_path is None:
        excel_path = _encontrar_excel_automatico()
    if excel_path is None:
        print("ERROR: No se encontro DatosPruebas2026_1.xlsx.", flush=True)
        print("  Usa --excel <ruta> para especificarlo manualmente.", flush=True)
        sys.exit(1)
    if not os.path.exists(excel_path):
        print(f"ERROR: El archivo no existe: {excel_path}", flush=True)
        sys.exit(1)

    t0 = time.perf_counter()
    resultados, ruta_excel = procesar_excel(
        excel_path=excel_path,
        max_pruebas=args.max_pruebas,
        solo_N=args.solo_N,
        sin_geo=args.sin_geo,
    )
    rA, rB, rC = generar_graficas(resultados)
    t_total = time.perf_counter() - t0

    imprimir_resumen(resultados, ruta_excel, rA, rB, rC)
    print(f"  Tiempo total de ejecucion: {t_total:.1f}s ({t_total/60:.1f} min)",
          flush=True)
    print()


if __name__ == "__main__":
    main()
